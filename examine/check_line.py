import openpyxl
from openpyxl.utils import get_column_letter

# エクセルファイルのパス
file_path = "C:/Users/1109685/Documents/IR/Sample_freejob_short.xlsx"

# ワークブックを開く
wb = openpyxl.load_workbook(file_path)
sheet = wb.active

def has_border(cell):
    """セルに枠線があるかどうかを確認する関数"""
    border = cell.border
    
    # 枠線のスタイルがあるかどうかを確認
    def is_bordered(side):
        return side is not None and side.style is not None and side.style != ""
    
    # 上下左右いずれかの辺に枠線があるかチェック
    return any([
        is_bordered(border.left),
        is_bordered(border.right),
        is_bordered(border.top),
        is_bordered(border.bottom)
    ])

# シートの使用範囲内のすべてのセルをチェック
for row in range(1, sheet.max_row + 1):
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=row, column=col)
        
        # セルの位置を特定（例：A1, B2など）
        cell_position = f"{get_column_letter(col)}{row}"
        
        # セルの値を取得（文字列に変換）
        cell_value = str(cell.value) if cell.value is not None else "空白"
        
        # 枠線があれば出力
        if has_border(cell):
            print(f"セル {cell_position} ({cell_value}): 枠線あり！")

print("チェック完了")
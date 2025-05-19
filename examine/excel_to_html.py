import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles.borders import Border

input_path = "C:/Users/1109685/Documents/IR/Sample_freejob_short.xlsx"
output_path = "output/data_output_color_short.html"

df = pd.read_excel(input_path)
wb = load_workbook(input_path)
sheet = wb.active

def has_border(cell):
    border: Border = cell.border
    def is_thick(side):
        return side is not None and side.style is not None and side.style != ""
    return any([
        is_thick(border.left),
        is_thick(border.right),
        is_thick(border.top),
        is_thick(border.bottom)
    ])

html = ['<table style="border-collapse: collapse; border: 1px solid black;">']

for r_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
    html.append("<tr>")
    for c_idx, cell in enumerate(row):
        if r_idx == 0:
            value = df.columns[c_idx] if c_idx < len(df.columns) else ""
        else:
            value = df.iat[r_idx - 1, c_idx] if (r_idx -1) < len(df.index) and c_idx < len(df.columns) else ""
        value = "" if pd.isna(value) else str(value)

        fill = cell.fill
        color = fill.fgColor.rgb if fill.fgColor.type == "rgb" else None
        rgb_style = f"background-color: #{color[-6:].lower()};" if color and color != "00000000" else ""

        base_border = "border: 1px solid black;"
        extra_border = "border: 5px solid black;" if has_border(cell) else ""

        style_str = f" style='{rgb_style} {extra_border or base_border}'"
        html.append(f"<td{style_str}>{value}</td>")
    html.append("</tr>")

html.append("</table>")

with open(output_path, "w", encoding="utf-8") as f:
    f.write('\n'.join(html))

print(f"色付きHTMLテーブルを保存しました: {output_path}")

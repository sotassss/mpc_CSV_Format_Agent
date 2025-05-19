import pandas as pd
from openpyxl import load_workbook

input_path = "C:/Users/1109685/Documents/IR/Sample_freejob_short.xlsx"
output_path = "output/data_output_color_short.md"

df = pd.read_excel(input_path)

# Markdownに変換（色付け前のデータ）
markdown_lines = df.to_markdown(index=False).split("\n")

# Markdownテーブルのデータ部分だけ分割
table_lines = [line for line in markdown_lines if '|' in line]
table = [line.strip('|').split('|') for line in table_lines]

# 背景色取得と適用
wb = load_workbook(input_path)
sheet = wb.active
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        # print(cell.coordinate)
        fill = cell.fill
        color = fill.fgColor.rgb if fill.fgColor.type == "rgb" else None

        if color != "00000000":
            print(f"{cell.coordinate} 色:{color}")
            r = cell.row -2 # pandasもmarkdownも0-indexed
            c = cell.column - 1
            try:
                original = table[r + 2][c].strip()  # +2: Markdownは見出しと区切り行がある
                # ARGBの末尾6桁を小文字にして取得
                rgb_lower = color[-6:].lower()
                colored = f"<span style='background-color: #{rgb_lower};'>{original}</span>"
                table[r + 2][c] = f" {colored} "  # スペースを入れてMarkdown整形を保つ
            except IndexError:
                continue

# 再構築
modified_lines = ['| ' + ' | '.join(row) + ' |' for row in table]

# 保存
with open(output_path, "w", encoding="utf-8") as f:
    f.write('\n'.join(modified_lines))

print("色付きMarkdownテーブルを保存しました:", output_path)

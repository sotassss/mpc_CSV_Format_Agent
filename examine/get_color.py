# エクセルの背景色を取得する
from openpyxl import load_workbook

wb = load_workbook("c:/Users/1109685/Documents/IR/Sample_freejob_short.xlsx")
sheet = wb.active

for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        # 背景色の取得
        fill = cell.fill
        color = fill.fgColor.rgb if fill.fgColor.type == "rgb" else None

        if color != "00000000":
            print(f"{cell.coordinate} 色:{color}")
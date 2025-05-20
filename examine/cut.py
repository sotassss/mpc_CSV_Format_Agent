import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy
import pandas as pd
import numpy as np

# 入力・出力ファイル
input_path = "C:/Users/1109685/Documents/IR/Sample_freejob_short.xlsx"
output_path = "C:/Users/1109685/Documents/IR/Sample_freejob_short_format.xlsx"

# Excel読み込み
wb_src = openpyxl.load_workbook(input_path)
ws_src = wb_src.active

# 値だけを一度DataFrame化（行列そのまま）
data = []
for row in ws_src.iter_rows(values_only=True):
    data.append(list(row))
df = pd.DataFrame(data)

# 空白・空文字をNaNに
df = df.replace(r'^\s*$', np.nan, regex=True)

# 「すべてがNaNまたは'nan'含む列」の検出と除去
def is_all_nan_or_contains_nan_string(col):
    for v in col:
        if pd.isna(v):
            continue
        if 'nan' not in str(v).lower():
            return False
    return True

cols_to_drop = [col for col in df.columns if is_all_nan_or_contains_nan_string(df[col])]
df_cleaned = df.drop(columns=cols_to_drop)

# 新規ブック作成
wb_new = Workbook()
ws_new = wb_new.active
ws_new.title = ws_src.title

# コピー対象列の元のインデックス（Excel列と一致させる）
retained_col_indices = [i for i in range(df.shape[1]) if i not in cols_to_drop]

# データと書式をコピー
for r in range(df_cleaned.shape[0]):
    for new_c, original_c in enumerate(retained_col_indices):
        src_cell = ws_src.cell(row=r + 1, column=original_c + 1)
        dst_cell = ws_new.cell(row=r + 1, column=new_c + 1)

        # 値をコピー
        dst_cell.value = src_cell.value

        # 書式をコピー
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.number_format = src_cell.number_format
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.protection = copy(src_cell.protection)

# 保存
wb_new.save(output_path)
print(f"保存完了：{output_path}")
